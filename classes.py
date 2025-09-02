from enum import Enum
import json
from copy import deepcopy

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill


class CellStatus(Enum):
    """BLANK, CROSS or QUEEN. Value is a string representing a suitable character.
    """
    BLANK = " "
    CROSS = "x"
    QUEEN = chr(0x2655) # ♕


class Cell:
    """A single cell of the Queens board
    """

    color: str
    """expected to be a 6-digit color value hex: e.g. 'FF0010'"""
    x: int
    """x-coordinate (0-indexing) \n\n x-coord is from left to right. y-coord is from top to bottom. Origin is at the top-left of the board."""
    y: int
    """y-coordinate (0-indexing) \n\n x-coord is from left to right. y-coord is from top to bottom. Origin is at the top-left of the board."""
    status: CellStatus

    def __init__(self, x: int, y: int, color: str):
        """x-coord is from left to right. y-coord is from top to bottom. Origin is at the top-left of the board.

        Args:
            color (str): 6-digit hex code. e.g. 'FF0010'
        """
        self.color = color; 
        self.x = x; self.y = y
        self.status = CellStatus.BLANK


class ColorSet:
    """A group of cells on the Queens board with the same color
    """

    color: str
    """expected to be a 6-digit color value hex: e.g. 'FF0010'"""
    _held_rows: set[int]
    """The row numbers held by the ColorSet's blank cells. 0-indexed. e.g. if the color set's blank cells spans the first 3 rows, this set would have 0,1,2"""
    _held_cols: set[int]
    """The column numbers held by the ColorSet's blank cells. 0-indexed. e.g. if the color set's blank cells spans the first 3 columns, this set would have 0,1,2"""
    cells: list[Cell]

    def __init__(self, cells: list[Cell], color: str):
        """_summary_

        Args:
            color (str): expected to be a 6-digit color value hex: e.g. 'FF0010'
        """

        self.color = color
        self._held_cols = set(); self._held_rows = set()

        self.cells = []
        for cell in cells:
            if cell.color == self.color: 
                self.cells.append(cell)
                self._held_cols.add(cell.x)
                self._held_rows.add(cell.y)

    def refresh_holdings(self):
        """Refresh the held_rows and cols of the object."""
        self._held_cols = set(); self._held_rows = set()
        for cell in self.cells:
            if cell.status == CellStatus.BLANK:
                self._held_cols.add(cell.x)
                self._held_rows.add(cell.y)

    def get_blank_cells(self) -> set[Cell]:
        blank_cells = set()
        for cell in self.cells:
            if cell.status == CellStatus.BLANK: blank_cells.add(cell)
        return blank_cells


class Board:
    """The Queens board.
    
    x-coord is from left to right. y-coord is from top to bottom. Origin is at the top-left of the board.
    """

    length: int
    """i.e. how many columns the board has
    """
    height: int
    """i.e. how many rows the board has
    """
    color_sets: dict[str, ColorSet]
    """Maps 6-digit color hexcode to ColorSets"""

    cell_grid: list[list[Cell]]
    """Convention such that we can access a cell via y,x.  i.e. self.cell_grid[y][x].\n 
    So the self.cell_grid would be a list of rows.
    """

    def __init__(self, length: int, height: int, cells: list[Cell]):
        self.length = length
        self.height = height
        self.color_sets = {}

        # initialize cell grid, so we can add the cells to it later in any order
        self.cell_grid = []
        for _ in range(0, height):
            row = [None] * length
            self.cell_grid.append(row)

        # add the cells to the grid
        for cell in cells:
            self.cell_grid[cell.y][cell.x] = cell

        # create the color sets
        same_colored_cells: dict[str, list[Cell]] = {}
        for cell in cells:
            if cell.color not in same_colored_cells.keys():
                same_colored_cells[cell.color] = [cell]
            else:
                same_colored_cells[cell.color].append(cell)

        for color in same_colored_cells.keys():
            color_set = ColorSet(
                cells=same_colored_cells[color],
                color=color
            )
            self.color_sets[color] = color_set
    
    @classmethod
    def from_json(cls, filepath: str):
        """Initialize the board from the json representing the board.
        """
        with open(filepath, "rt") as f:
            data = json.load(f)

        # the json has the cell color values as a list of rows.
        # Each row displays the values from left to right
        # The rows are listed in the order top to bottom.
        
        cells: list[Cell] = []

        cell_data: list[list[str]] = data["colors"]
        for i, row in enumerate(cell_data):
            for j, color in enumerate(row):
                if color == 'SystemButtonFace': color = "#FFFFFF" # if this is the case assume white
                cells.append(Cell(j, i, color[1:])) # remove the leading '#' from the otherwise 6-digit color hex code

        return cls(
            length = int(data["cols"]),
            height = int(data["rows"]),
            cells = cells
        )



    def to_excel(self, filepath: str, offset: int = 0):
        """Save the board (in its current state) to an excel file for visualization.

        If offset is 0, creates new/overwrites excel file. Else adds to the existing excel file.
        This is done so that incremental stages of solving of the board can be printed vertically on the same excel file.

        Args:
            filepath (str): _description_
            offset (int, optional): How many rows to skip ahead when printing the board to excel cells. If offset is 0, creates new/overwrites excel file. Else adds to the existing excel file. Defaults to 0.
        """
        if offset == 0:
            wb = Workbook()
        else:
            wb = load_workbook(filepath)
        ws = wb.active

        offset_rows = (offset * self.height) + offset

        # set column and row lengths (note: openpyxl uses 1-indexing. It also takes columns as left to right and rows as top to bottom.)
        for i in range(1, self.height + 1):
            ws.row_dimensions[i + offset_rows].height = 20
        for i in range(1, self.length + 1):
            ws.column_dimensions[get_column_letter(i)].width = 4

        for row in self.cell_grid:
            for cell in row:
                excel_cell = ws.cell(row=cell.y + 1 + offset_rows, column=cell.x + 1, value=cell.status.value) # openpyxl will use 1-indexing
                excel_cell.fill = PatternFill(start_color=cell.color, fill_type="solid")

        wb.save(filepath)

    def to_status_grid(self) -> list[list[str]]:
        """Save the statuses of the board's cells to a list of lists and save as a pickle file.

        Used for testing purposes.
        An example of the list of lists:
        [
            [' ', 'x', '♕'],
            ['x', 'x', ' '],
            ['x', 'x', ' ']
        ]

        Args:
            filepath (str): filepath without the '.pkl' extension

        Returns:
            list[list[str]]: List of lists (rows) containing the cell statuses (see the function summary for an example)
        """
        grid = []
        for row in self.cell_grid:
            row_statuses = []
            for cell in row:
                row_statuses.append(cell.status.value)
            grid.append(row_statuses)
        return grid


    def __would_block_cells(self, cell: Cell) -> set[Cell]:
        """If the input cell is made queen, get the set of cells that it would block.
        """
        x = cell.x; y = cell.y

        blocked_cells: set[Cell] = set()

        # cells on the same row and column
        for col_x in range(0, self.length):
            cell_ = self.cell_grid[y][col_x]
            blocked_cells.add(cell_)
        for row_y in range(0, self.height):
            cell_ = self.cell_grid[row_y][x]
            blocked_cells.add(cell_)

        # adjacent cells (diagonal)
        for col_x in [x-1, x+1]:
            for row_y in [y-1, y+1]:
                if col_x < 0 or row_y < 0: continue # normally this would trigger an IndexError and safely skip but this is Python and it uses negative indexing so...
                try: 
                    cell_ = self.cell_grid[row_y][col_x]
                    blocked_cells.add(cell_)
                except IndexError: continue

        # cells of the same color
        same_color_set = self.color_sets[cell.color]
        for cell_ in same_color_set.cells:
            blocked_cells.add(cell_)

        # except itself ofcourse
        blocked_cells.remove(cell)

        return blocked_cells
    
    def get_cell_at(self, x: int, y: int) -> Cell:
        return self.cell_grid[y][x]


    def __mark_queen(self, cell: Cell):
        """Mark a cell as a queen. Cross out all the blank cells it would block
        """
        x = cell.x; y = cell.y

        # mark cell as queen
        self.cell_grid[y][x].status = CellStatus.QUEEN

        # cross out all the blank cells it would block
        would_block_cells = self.__would_block_cells(cell)
        for cell in would_block_cells:
            if cell.status == CellStatus.BLANK: cell.status = CellStatus.CROSS


    def mark_queens_where_certain(self) -> bool:
        """Mark queens on the board where certain.

        Returns:
            bool: Returns true if at least one queen was marked.
        """
        queen_marked = False

        # if a row only has one blank cell
        blank_cells: list[Cell] = []
        for row in self.cell_grid:
            for cell in row:
                if cell.status == CellStatus.BLANK: blank_cells.append(cell)
            if len(blank_cells) == 1:
                # print("row only has one blank cell")
                cell = blank_cells[0]
                self.__mark_queen(cell)
                queen_marked = True
            blank_cells = [] # reset for next row

        # if a column only has one blank cell
        blank_cells: list[Cell] = []
        for col_x in range(0, self.length):
            for row_y in range(0, self.height):
                cell = self.cell_grid[row_y][col_x]
                if cell.status == CellStatus.BLANK: blank_cells.append(cell)
            if len(blank_cells) == 1:
                # print("col only has one blank cell")
                cell = blank_cells[0]
                self.__mark_queen(cell)
                queen_marked = True
            blank_cells = [] # reset for next column

        # if a color set only has one blank cell
        blank_cells: list[Cell] = []
        for color_set in self.color_sets.values():
            for cell in color_set.cells:
                if cell.status == CellStatus.BLANK: blank_cells.append(cell)
            if len(blank_cells) == 1:
                # print("color set only has one blank cell")
                cell = blank_cells[0]
                self.__mark_queen(cell)
                queen_marked = True
            blank_cells = [] # reset for next color set

        return queen_marked

    
    def __refresh_color_set_holdings(self):
        """Refresh the held_rows and cols of the color sets in the board."""
        for color_set in self.color_sets.values(): color_set.refresh_holdings()

    @staticmethod
    def has_board_changed(board_1: 'Board', board_2: 'Board') -> bool:
        """Checks if statuses of cells (blank, crossed, queen) are the same between the 2 boards.
        """
        for col_x in range(0, board_1.length):
            for row_y in range(0, board_1.height):
                cell_1 = board_1.cell_grid[row_y][col_x]
                cell_2 = board_2.cell_grid[row_y][col_x]
                if cell_1.status != cell_2.status: return True
        return False
    
    def is_game_over(self) -> bool:
        """If all the Queens have been found. This method only makes sense for a square board...  I mean if it wasn't it would be a very abnormal queens game.
        """
        queen_count = 0
        for row in self.cell_grid:
            for cell in row:
                if cell.status == CellStatus.QUEEN: queen_count += 1

        if queen_count == self.height: return True
        return False
    
    def would_cell_block_color_set(self, cell: Cell) -> bool:
        """Assuming a cell is a Queen, would it block any other color set completely?
        """
        would_block_cells = self.__would_block_cells(cell)

        # for each color set, get a list of blank cells. If all blank cells are included in would_block_cells, that color set would be blocked.
        for color_set in self.color_sets.values():
            if color_set.color == cell.color: continue

            blank_cells = color_set.get_blank_cells()
            if len(blank_cells) == 0: continue
            if blank_cells.issubset(would_block_cells): return True

        return False
    
    def would_cell_block_color_set_n(self, cell: Cell, n: int) -> bool:
        """Assuming a cell is a Queen, would it block any other color set completely?

        Uses recursion to check n number of moves ahead.
        For example it might be required to check 2 moves ahead:
        There can be a cell that would not block any other colorset completely by itself, however marking any of the remaining cells that would be left after it
        is marked queen would result in a colorset getting blocked.

        Args:
            cell (Cell): _description_
            n (int): How many moves to check ahead.

        Returns:
            bool: _description_
        """
        # if n == 0: return False
        if n == 1: return self.would_cell_block_color_set(cell) # optimization, avoids unnecessary deepcopy below. It also makes the above line obsolete.

        # copy the board
        copy_board = deepcopy(self)
        the_cell = copy_board.get_cell_at(cell.x, cell.y)

        if copy_board.would_cell_block_color_set(the_cell): return True
        
        copy_board.__mark_queen(the_cell)

        blank_cells = copy_board.get_blank_cells()
        results: list[bool] = []
        for _cell in blank_cells:
            result = copy_board.would_cell_block_color_set_n(_cell, n-1) # recursion
            results.append(result)

        # if at least one result is True, return True
        for result in results:
            if result: return True

        return False


    def get_blank_cells(self) -> list[Cell]:
        blank_cells = []
        for row in self.cell_grid:
            for cell in row:
                if cell.status == CellStatus.BLANK: blank_cells.append(cell)
        return blank_cells
        
    
    def colorset_axis_holdings(self, axis: str) -> dict[str, frozenset[int]]:
        """Returns dictionary mapping color to the set of rows or columns held by it.

        Args:
            axis (str): 'row' or 'col'

        Raises:
            Exception: If you enter something other than 'row' or 'col' for the axis argument.

        Returns:
            dict[str, frozenset[int]]: Dictionary mapping color to the set of rows or columns held by it.
        """
        self.__refresh_color_set_holdings()

        colorset_axis_holdings = {}
        if axis == 'row': held = "_held_rows"
        elif axis == 'col': held = "_held_cols"
        else: raise Exception()

        for colorset in self.color_sets.values():
            colorset_axis_holdings[colorset.color] = frozenset(colorset.__getattribute__(held))

        return colorset_axis_holdings

